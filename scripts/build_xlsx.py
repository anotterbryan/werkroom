#!/usr/bin/env python3
"""
build_xlsx.py  --  Compile data/*.csv into dist/DragRaceDB_master.xlsx.

A faithful, human-readable mirror of the source CSVs: one worksheet per table plus a
leading "index" sheet. Numeric columns are written as numbers (so sorting/filtering
works); IDs and text stay text; unknowns stay "-" per the project's no-guessing rule.

Usage:
    python scripts/build_xlsx.py
"""
import csv, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "dist" / "DragRaceDB_master.xlsx"
BLANKS = ("", "-")

# table order + the columns that should be written as integers when numeric
TABLES = ["queens", "seasons", "contestants", "episodes", "progression",
          "songs", "lip_syncs", "elimination_events", "panel", "appearances"]
NUMERIC = {"placement", "entrance_order", "wins", "highs", "lows", "bottoms",
           "episode_number", "episode_count", "times_used"}

HEADER_FILL = PatternFill("solid", fgColor="2A0E3A")   # pageant-noir deep purple
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)


def read(name):
    with open(DATA / f"{name}.csv", newline="", encoding="utf-8") as fh:
        r = csv.DictReader(fh)
        return r.fieldnames, list(r)


def style_sheet(ws, headers, nrows):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if nrows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{nrows + 1}"
    # column widths from header + sampled content
    for c, h in enumerate(headers, 1):
        width = max(len(h) + 4, 10)
        for row in ws.iter_rows(min_row=2, max_row=min(nrows + 1, 60),
                                min_col=c, max_col=c):
            v = row[0].value
            if v is not None:
                width = max(width, min(len(str(v)) + 2, 48))
        ws.column_dimensions[get_column_letter(c)].width = width


def main():
    wb = Workbook()
    idx = wb.active
    idx.title = "index"

    counts = []
    for name in TABLES:
        headers, rows = read(name)
        ws = wb.create_sheet(name)
        for r, row in enumerate(rows, 2):
            for c, h in enumerate(headers, 1):
                v = row[h]
                if h in NUMERIC and str(v).strip().lstrip("-").isdigit():
                    ws.cell(row=r, column=c, value=int(v)).font = BODY_FONT
                else:
                    ws.cell(row=r, column=c, value=v).font = BODY_FONT
        style_sheet(ws, headers, len(rows))
        counts.append((name, len(rows), len(headers)))

    # index sheet
    idx.cell(row=1, column=1, value="RPDR Tracking - master workbook").font = \
        Font(name="Arial", size=14, bold=True, color="E6007E")
    idx.cell(row=2, column=1,
             value=f"Generated {datetime.date.today().isoformat()} from data/*.csv "
                   "(the source of truth). Do not hand-edit this file; edit the CSVs "
                   "and re-run scripts/build_xlsx.py.").font = BODY_FONT
    hdr = ["table", "rows", "columns"]
    for c, h in enumerate(hdr, 1):
        cell = idx.cell(row=4, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, (name, nr, nc) in enumerate(counts, 5):
        idx.cell(row=i, column=1, value=name).font = BODY_FONT
        idx.cell(row=i, column=2, value=nr).font = BODY_FONT
        idx.cell(row=i, column=3, value=nc).font = BODY_FONT
    idx.column_dimensions["A"].width = 24
    idx.column_dimensions["B"].width = 10
    idx.column_dimensions["C"].width = 10
    idx.freeze_panes = "A5"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)} with {len(TABLES)} table sheets + index")


if __name__ == "__main__":
    main()
